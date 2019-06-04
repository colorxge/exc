import openpyxl
ac = openpyxl.Workbook()
# 导入文件对象
xlsx_fp = openpyxl.load_workbook('t2.xlsx')
# 创建新的表格
xlsx_fp.create_sheet('刀剑如梦', 1)
print(xlsx_fp.sheetnames)
print(xlsx_fp['同城交友'])
# 删除表格
a = xlsx_fp['同城交友']
# xlsx_fp.remove(a)
print(xlsx_fp.sheetnames)

# 查看表格属性
print(a.title, a.max_row, a.min_column)
print(a.__dict__.keys())
# 改变活跃的表格
xlsx_fp.active = a
print(xlsx_fp.active)
print(a.merged_cells)
# 合并单元格
a.merge_cells( start_row=1, start_column=1, end_row=3, end_column=3)
# xlsx_fp.save('t2.xlsx')